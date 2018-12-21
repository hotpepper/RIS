__author__ = 'SHostetter'
import os
import csv
import datetime
from xlrd import open_workbook, xldate_as_tuple  # xls
from openpyxl import load_workbook  # xlsx
from Tkinter import Tk
from tkFileDialog import askopenfilename
import tkMessageBox


class CSVDataFile:
    def __init__(self, input_file):
        self.file_name = os.path.splitext(input_file)[0]
        self.file_type = os.path.splitext(input_file)[1]
        self.file_path = os.path.dirname(os.path.abspath(input_file))
        self.columns = 0
        self.rows = 0
        self.headers = list()  # will be first row of data
        self.data = list()  # list of dictionaries of tab or csv data
        self.sheet_name = self.file_name.split('/')[-1]

    def dict_line(self, data_list):
        """
        :param self.headers: list of columns / headers
        :param data_list: data row (list format)
        :return: dicitonary mapping the row data to the column headers
        """
        d = dict()
        for c in range(len(self.headers)):
            d[self.headers[c]] = data_list[c]
        return d

    def check_for_dates(self, cv):
        tests = (lambda value: datetime.datetime.strptime(value, '%m/%d/%Y'),
                 lambda value: datetime.datetime.strptime(value, '%m/%d/%y')
                 )
        try:
            return tests[0](cv)
        except:
            try:
                return tests[1](cv)
            except:
                return cv

    def read(self):
        csv_file = os.path.join(self.file_path, self.file_name+'.csv')
        # mimics the csv.DictReader
        with open(csv_file, 'rb') as f:
            for row in csv.reader(f):
                if not self.rows:
                    self.headers = [str(i).upper() for i in row]
                    self.columns = len(row)
                    self.rows += 1
                else:  # data
                    row = [self.check_for_dates(c) for c in row]
                    self.data.append(self.dict_line(row))
                    self.rows += 1
        print '%i rows read from file %s%s' % (self.rows, self.file_name, self.file_type)


class XLSDataFile:
    """
    for excel documents each tab should be passed as seperate object
    """
    def __init__(self, input_file, tab=None):
        self.file_name = os.path.splitext(input_file)[0]
        self.file_type = os.path.splitext(input_file)[1]
        self.file_path = os.path.dirname(os.path.abspath(input_file))
        self.book = open_workbook(input_file)
        self.sheet_name = tab
        self.tab = self.book.sheet_by_name(tab)
        self.columns = self.tab.ncols
        self.rows = self.tab.nrows
        self.headers = list()  # will be first row of data
        self.data = list()  # list of dictionaries of tab or csv data

    def dict_line(self, data_list):
        """
        :param self.headers: list of columns / headers
        :param data_list: data row (list format)
        :return: dicitonary mapping the row data to the column headers
        """
        d = dict()
        for c in range(len(self.headers)):
            d[self.headers[c]] = data_list[c]
        return d

    def cell_value(self, sheet, row, column):
        """
        XL_CELL_EMPTY	0 empty string
        XL_CELL_TEXT	1 a Unicode string
        XL_CELL_NUMBER	2 float
        XL_CELL_DATE	3 float
        XL_CELL_BOOLEAN	4 int; 1 means True, 0 means False
        XL_CELL_ERROR	5
        XL_CELL_BLANK	6
        :param sheet:
        :param row:
        :param column:
        :return:
        """
        if sheet.cell(row, column).value:
            if sheet.cell(row, column).value != u'\x00':
                if sheet.cell(row, column).ctype == 3:  # Date
                    date_value = xldate_as_tuple(sheet.cell(row, column).value, self.book.datemode)
                    return datetime.datetime(*date_value)
                elif type(sheet.cell(row, column).value) in (float, int):
                    return float(sheet.cell(row, column).value)
                elif sheet.cell(row, column).value not in (u'\xc2', u'\xa0'):
                    return sheet.cell(row, column).value.encode('utf-8').strip().replace('\xe2\x80\x99', "")

    def read(self):
        for r in range(self.rows):
            row = [self.cell_value(self.tab, r, c) for c in range(self.columns)]
            if r == 0:
                self.headers = [str(i).upper() for i in row]
            else:
                self.data.append(self.dict_line(row))
        print '%i rows read from file %s%s tab (%s)' % (self.rows, self.file_name, self.file_type, self.sheet_name)


class XLSXDataFile:
    """
    for excel documents each tab should be passed as seperate object
    """
    def __init__(self, input_file, tab=None):
        self.file_name = os.path.splitext(input_file)[0]
        self.file_type = os.path.splitext(input_file)[1]
        self.file_path = os.path.dirname(os.path.abspath(input_file))
        self.book = load_workbook(input_file, data_only=True, read_only=True)
        self.sheet_name = tab
        self.tab = self.book[tab]
        self.columns = self.tab.max_column
        self.rows = self.tab.max_row
        self.headers = list()  # will be first row of data
        self.data = list()  # list of dictionaries of tab or csv data

    def dict_line(self, data_list):
        """
        :param self.headers: list of columns / headers
        :param data_list: data row (list format)
        :return: dicitonary mapping the row data to the column headers
        """
        d = dict()
        for c in range(len(self.headers)):
            d[self.headers[c]] = data_list[c]
        return d

    def read(self):
        for row in self.tab.rows:
            if not self.headers:
                for c in row:
                    self.headers.append(str(c.value).upper())
            else:
                line = list()
                for c in row:
                    if type(c.value) in (unicode, str):
                        v = c.value.replace(u'\xa0', u' ')
                        line.append(v)
                    # print c.value, type(c.value)
                    else:
                        line.append(c.value)
                self.data.append(self.dict_line(line))
        print '%i rows read from file %s%s tab (%s)' % (self.rows, self.file_name, self.file_type, self.sheet_name)


class DataFile:
    def __init__(self, file_path=None):
        self.file = file_path
        self.file_loc()
        self.data = dict()
        self.controller()
        self.intersections = self.tabs_with_intersection_min()
        self.stretches = self.tabs_with_stretch_min()
        self.addresses = self.tabs_with_addresses_min()
        self.parse_address()
        self.nodes = self.tabs_with_nodes_min()

    def file_loc(self):
        if not self.file:
            Tk().withdraw()
            tkMessageBox.showinfo("Open file", "Please navigate to the Excel or CSV file you want to process")
            filename = askopenfilename()
            self.file = filename

    def controller(self):
        print 'Processing %s...\n' % self.file
        file_type = os.path.splitext(self.file)[1]
        if file_type == '.csv':
            # get name for dict key (proxy for excel tabs)
            name = os.path.splitext(self.file)[0]
            name = name.split(r'/')[-1]
            self.data[name] = CSVDataFile(self.file)
            self.data[name].read()
            # print self.data[name].headers
        if file_type == '.xls':
            # get worksheets to add to dictionary
            sheets = open_workbook(self.file).sheet_names()
            name = os.path.splitext(self.file)[0]
            name = name.split(r'/')[-1]
            for tab in sheets:
                sheet_name = name+'_'+tab
                self.data[sheet_name] = XLSDataFile(self.file, tab)
                self.data[sheet_name].read()
                # print sheet_name, self.data[sheet_name].headers
        if file_type == '.xlsx':
            # get worksheets to add to dictionary
            sheets = load_workbook(self.file, data_only=True).sheetnames
            name = os.path.splitext(self.file)[0]
            name = name.split(r'/')[-1]
            for tab in sheets:
                sheet_name = name+'_'+tab
                self.data[sheet_name] = XLSXDataFile(self.file, tab)
                self.data[sheet_name].read()
                # print sheet_name, self.data[sheet_name].headers

    def tabs_with_intersection_min(self):
        ints = list()
        for tab in self.data.keys():
            if 'STREET 1' in self.data[tab].headers:
                if 'STREET 2' in self.data[tab].headers and 'STREET 3' not in self.data[tab].headers:
                    if 'BOROUGH' in self.data[tab].headers:
                        ints.append(tab)
        return ints

    def tabs_with_stretch_min(self):
        stretch = list()
        for tab in self.data.keys():
            if set(['STREET 1', 'STREET 2', 'STREET 3', 'BOROUGH']).issubset(set(self.data[tab].headers)):
                stretch.append(tab)
            elif set(['ON', 'FROM', 'TO', 'BOROUGH']).issubset(set(self.data[tab].headers)):
                stretch.append(tab)
            elif set(['ON STREET', 'FROM STREET', 'TO STREET', 'BOROUGH']).issubset(set(self.data[tab].headers)):
                stretch.append(tab)
        return stretch

    def tabs_with_addresses_min(self):
        addrs = list()
        for tab in self.data.keys():
            if 'NUMBER' in self.data[tab].headers:
                if 'STREET' in self.data[tab].headers:
                    if 'BOROUGH' in self.data[tab].headers:
                                addrs.append(tab)
        return addrs

    def tabs_with_nodes_min(self):
        nodes = list()
        for tab in self.data.keys():
            if 'NODEID' in self.data[tab].headers:
                nodes.append(tab)
        return nodes

    def parse_address(self):
        # if not self.tabs_with_addresses_min():
        for tab in self.data.keys():
            if 'ADDRESS' in self.data[tab].headers:
                for row in self.data[tab].data:
                    address = row['ADDRESS']
                    if address:
                        row['NUMBER'] = address.split()[0]
                        row['STREET'] = address.replace(address.split()[0], '').lstrip()
                        if 'NUMBER' not in self.data[tab].headers:
                            self.data[tab].headers.append('NUMBER')
                        if 'STREET' not in self.data[tab].headers:
                            self.data[tab].headers.append('STREET')
                    else:
                        # add keys for skipped rows
                        if 'NUMBER' in self.data[tab].headers:
                            row['NUMBER'] = '0'
                            row['STREET'] = ''
        self.addresses = self.tabs_with_addresses_min()



if __name__ == '__main__':
    test_csv = DataFile('C:/Users/SHostetter/Desktop/GIT/RIS/geocoder/test_files/test.csv')
    print test_csv.intersections
    test_xls = DataFile('C:/Users/SHostetter/Desktop/GIT/RIS/geocoder/test_files/test.xls')
    print test_xls.intersections
    print test_xls.addresses
    test_xlsx = DataFile('C:/Users/SHostetter/Desktop/GIT/RIS/geocoder/test_files/test.xlsx')
    print test_xlsx.intersections
    # test = DataFile('C:/Users/SHostetter/Desktop/GIT/RIS/geocoder/test_files/dt.csv')
    test = DataFile()
    print test.intersections
    
